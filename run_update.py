import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import copy

# 自動抓今天的年月
month_tag = datetime.today().strftime("%Y%m")
report_file = f"IM/{month_tag}_HL_Maintain_Report.xlsx"
data_file = "data.xlsx"

# 讀取本地資料庫
data_wb = openpyxl.load_workbook(data_file)
data_ws = data_wb["IM"]

# 找出台芝工作案號欄位
def find_case_column(ws):
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and "工作案號" in str(cell.value):
                return cell.column
    return None

# 找出時間欄位 index
def find_x_col_index(ws):
    for row in ws.iter_rows(min_row=1, max_row=5):
        for idx, cell in enumerate(row):
            if cell.value and "時間" in str(cell.value):
                return idx
    return None

# 找出台芝案號欄
case_col = find_case_column(data_ws)
if not case_col:
    raise ValueError("❌ 無法在 IM 分頁中找到台芝工作案號欄位")

# 找出最後一筆案號
last_case_id = None
for row in reversed(list(data_ws.iter_rows(min_row=2, values_only=True))):
    if row[case_col - 1]:
        last_case_id = str(row[case_col - 1]).strip()
        break

print("最後更新的案號為:", last_case_id)

# 讀取報表
report_wb = openpyxl.load_workbook(report_file)
report_ws = report_wb.active

# 找出報表中的案號與時間欄
report_case_col = find_case_column(report_ws)
x_col_index = find_x_col_index(report_ws)

# 收集新資料
start_append = False
append_rows = []
for row in report_ws.iter_rows(min_row=3):
    case_id = str(row[report_case_col - 1].value).strip() if row[report_case_col - 1] else ""
    if last_case_id and case_id == last_case_id:
        start_append = True
        continue
    if start_append:
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            break
        append_rows.append(values)

print(f"共 {len(append_rows)} 列新資料將加入 IM")

# 參考最後一列的樣式與公式（列高、格式）
ref_row = data_ws.max_row
ref_row_height = 21.66
ref_cells = {cell.column: cell for cell in data_ws[ref_row]}
al_formula_cell = data_ws[f"AL{ref_row}"]
am_formula_cell = data_ws[f"AM{ref_row}"]

# 寫入新資料
for row_data in append_rows:
    data_ws.append(row_data)
    new_row = data_ws.max_row
    data_ws.row_dimensions[new_row].height = ref_row_height

    for col_idx, value in enumerate(row_data, start=1):
        cell = data_ws.cell(row=new_row, column=col_idx)
        ref_cell = ref_cells.get(col_idx)
        if ref_cell:
            cell.font = copy.copy(ref_cell.font)
            cell.alignment = copy.copy(ref_cell.alignment)
            cell.border = copy.copy(ref_cell.border)
            cell.fill = copy.copy(ref_cell.fill)

    # 處理 X 欄時間格式
    if x_col_index is not None and x_col_index < len(row_data):
        x_value = row_data[x_col_index]
        try:
            if isinstance(x_value, str):
                dt = datetime.strptime(x_value.strip(), "%Y-%m-%d %H:%M:%S")
            elif isinstance(x_value, datetime):
                dt = x_value
            else:
                dt = None

            if dt:
                x_col_letter = get_column_letter(x_col_index + 1)
                cell = data_ws[f"{x_col_letter}{new_row}"]
                cell.value = dt
                cell.number_format = 'yyyy/m/d h:mm:ss AM/PM'
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        except Exception as e:
            print(f"❗ 日期格式轉換錯誤：{e}")

    # 複製公式到 AL 與 AM
    for col_letter, formula_cell in [("AL", al_formula_cell), ("AM", am_formula_cell)]:
        target_cell = data_ws[f"{col_letter}{new_row}"]
        formula = formula_cell.value.replace(str(ref_row), str(new_row)) if formula_cell.data_type == "f" else formula_cell.value
        target_cell.value = formula
        target_cell.font = copy.copy(formula_cell.font)
        target_cell.alignment = copy.copy(formula_cell.alignment)
        target_cell.border = copy.copy(formula_cell.border)
        target_cell.number_format = formula_cell.number_format

# 儲存檔案
data_wb.save(data_file)
print("✅ 資料更新完成")
